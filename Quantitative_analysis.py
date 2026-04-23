"""
data_loader.py
==============
Carga de fondos y precios con fallback en caso de que falten observaciones fuentes en cascada:

  1. SQL Server (SIIFSK01)  — fuente primaria para todos los fondos
  2. Fallback local (Excel) — para fondos de la lista FONDOS_CON_FALLBACK_LOCAL,
                              busca un archivo .xlsx en RUTA_FALLBACK_LOCAL cuyo
                              nombre coincide con el fondo (sin espacios).

Flujo de entrada:
  • El usuario carga un Excel donde:
      A1  = fondo objetivo
      A2+ = fondos candidatos preseleccionados
  • Se exige que TODOS los fondos del Excel tengan precios.
    Si falta alguno se lanza FondosFaltantesError antes de continuar.

Convención de Emision_Id:
  • Internamente se compara SIEMPRE sin espacios.
  • Se conserva el nombre original del Excel para mostrarlo en pantalla.
  • La BD normaliza con REPLACE(Emision_Id,' ','') en el WHERE.
"""

from __future__ import annotations
import os
import pandas as pd
from datetime import date
from pathlib import Path
from typing import Optional
import pyodbc
import pandas as pd
import numpy as np
from datetime import date
from datetime import timedelta
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from pathlib import Path
import re

# ----------------------------- Constantes ------------------------------------

# Fondos que NO están en la BD y tienen archivo Excel propio en la carpeta de red.
# La comparación se hace sin espacios.
FONDOS_CON_FALLBACK_LOCAL: set[str] = {
    "SURRVMXBFE",
    "SURUDIBFE",
    "SURPATBFE",
    "SUR30EBFE",
    "SURCTEBFE",
    "SURGOBBFE",
    "SURVEURBFE",
    "SUR-RFIBFE",
    "SK-GUBE",
    "SURASIABFE",
    "FT-CORPBE2",
}

RUTA_FALLBACK_LOCAL = Path(
    r"\\mxmefs02\Mesa de Operaciones\AAAA DENILSON\Fondo Estratega\Fallback_rendimientos"
)

#----------------------------- Funciones ------------------------------------

def normalizar_fondo(x):
    return str(x).strip().replace(" ", "")

def Consulta(Server:str,Database:str,Query:str):
    """
    Transforms a SQL query into a pandas data frame.
    
    Parameters
    ----------
    Server : Name of the server on SQL.
    Database : Name of the database.
    Query : String of the query that will be executed.

    Returns
    -------
    df : Returns optional data frame with the data extracted from the query.

    """
    df=None
    connection=None
    try:
        connection=pyodbc.connect(f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={Server};DATABASE={Database};Trusted_Connection=yes;')
        print("Conexión exitosa")
        cursor=connection.cursor()
        cursor.execute(Query)
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()
        df = pd.DataFrame.from_records(rows, columns=columns)
        #print(df.info())
        return df
    except Exception as ex:
        print(ex)
        return None
    finally:
        if connection is not None:
            connection.close()
            print("Conexión finalizada.")

def leer_fondos_excel(archivo) -> tuple[str, list[str], str | None]:
    """
    Lee el Excel de entrada:
        A1  = fondo objetivo
        A2+ = candidatos preseleccionados

    'archivo' es una ruta str/Path.

    Retorna (fondo_objetivo, candidatos, error_message).
    error_message es None si todo está bien.
    """
    try:
        df = pd.read_excel(archivo, header=None, usecols=[0], dtype=str)
    except Exception as e:
        return "", [], f"No se pudo leer el archivo de fondos: {e}"

    nombres = df[0].dropna().str.strip().tolist()
    nombres = [n for n in nombres if n]

    if len(nombres) < 1:
        return "", [], "El archivo está vacío. Debe tener al menos un fondo en A1."
    if len(nombres) < 2:
        return "", [], "El archivo solo tiene fondo objetivo (A1). Agrega al menos un candidato en A2."

    return nombres[0], nombres[1:], None

def formatear_lista_sql(lista):
    return ",".join([f"'{x.strip().replace(' ', '')}'" for x in lista])

def calcular_resumen_fondos(df, fondo_objetivo, dias_anualizacion=252):
    df = df.copy()

    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df["Rendimiento"] = pd.to_numeric(df["Rendimiento"], errors="coerce")
    df["Fondo_norm"] = df["Fondo"].apply(normalizar_fondo)

    fondo_obj_norm = normalizar_fondo(fondo_objetivo)

    # Serie del fondo objetivo
    df_obj = (
        df[df["Fondo_norm"] == fondo_obj_norm][["Fecha", "Rendimiento"]]
        .dropna()
        .rename(columns={"Rendimiento": "Rendimiento_Objetivo"})
        .sort_values("Fecha")
    )

    if df_obj.empty:
        raise ValueError(f"No se encontró información del fondo objetivo: {fondo_objetivo}")

    resultados = []

    for fondo_norm, grupo in df.groupby("Fondo_norm"):
        nombre_fondo = grupo["Fondo"].iloc[0]

        serie = (
            grupo[["Fecha", "Rendimiento"]]
            .dropna()
            .sort_values("Fecha")
            .copy()
        )

        if len(serie) < 2:
            continue

        rend = serie["Rendimiento"].to_numpy()

        # ===== Métricas standalone del fondo =====
        rendimiento_anual_promedio = (np.prod(1 + rend) ** (dias_anualizacion / len(rend))) - 1
        volatilidad_diaria = np.std(rend, ddof=1)
        volatilidad_anualizada = volatilidad_diaria * np.sqrt(dias_anualizacion)
        var_historico_95 = np.quantile(rend, 0.05)

        # Maximum Drawdown
        curva_acumulada = np.cumprod(1 + rend)
        maximo_acumulado = np.maximum.accumulate(curva_acumulada)
        drawdowns = (curva_acumulada / maximo_acumulado) - 1
        maximum_drawdown = np.min(drawdowns)

        fila = {
            "Fondo": nombre_fondo,
            "Rendimiento_anual_promedio": rendimiento_anual_promedio,
            "Volatilidad_diaria": volatilidad_diaria,
            "Volatilidad_anualizada": volatilidad_anualizada,
            "VaR_Historico_95": var_historico_95,
            "Maximum_Drawdown": maximum_drawdown,
            "R2": np.nan,
            "Beta": np.nan,
            "Alfa Anualizada": np.nan,
            "Tracking Error Residual Anual": np.nan,
            "Tracking Error Anual": np.nan,
            "Information Ratio": np.nan,
        }

        # ===== Regresión solo para fondos distintos al objetivo =====
        if fondo_norm != fondo_obj_norm:
            base = df_obj.merge(
                serie,
                on="Fecha",
                how="inner"
            )

            if len(base) >= 2:
                x = base["Rendimiento_Objetivo"].to_numpy()
                y = base["Rendimiento"].to_numpy()

                # y = alfa + beta * x
                beta, alfa = np.polyfit(x, y, 1)
                y_hat = alfa + beta * x
                error = y - y_hat
                alfa_anual = alfa * dias_anualizacion

                sse = np.sum(error ** 2)
                sst = np.sum((y - np.mean(y)) ** 2)
                r2 = 1 - (sse / sst) if sst != 0 else np.nan

                tracking_error_residual = np.std(error, ddof=1)
                te_residual_anual = tracking_error_residual * np.sqrt(dias_anualizacion)

                active_returns = y - x
                active_return_mean = np.mean(active_returns)
                tracking_error = np.std(active_returns, ddof=1)
                tracking_error_anual = tracking_error * np.sqrt(dias_anualizacion)

                information_ratio = (
                    (active_return_mean / tracking_error) * np.sqrt(dias_anualizacion)
                    if tracking_error != 0 else np.nan
                )

                fila.update({
                    "R2": r2,
                    "Beta": beta,
                    "Alfa Anualizada": alfa_anual,
                    "Tracking Error Residual Anual": te_residual_anual,
                    "Tracking Error Anual": tracking_error_anual,
                    "Information Ratio": information_ratio,
                })

        resultados.append(fila)

    resumen = pd.DataFrame(resultados)

    if resumen.empty:
        return resumen

    # Orden opcional: objetivo primero
    resumen["Es_Objetivo"] = resumen["Fondo"].apply(
        lambda x: normalizar_fondo(x) == fondo_obj_norm
    )
    resumen = resumen.sort_values(
        ["Es_Objetivo", "Fondo"],
        ascending=[False, True]
    ).drop(columns=["Es_Objetivo"])

    return resumen

def agregar_score_sustitucion_absoluto(
    df_resumen,
    fondo_objetivo,
    pesos=None,
    columna_comision="Comisión de Administración",
    umbral_te=0.10,
    umbral_ir=1.0,
    tolerancia_beta=0.30
):
    df = df_resumen.copy()

    if pesos is None:
        pesos = {
            "R2": 0.30,
            "Tracking Error Anual": 0.20,
            "Beta": 0.15,
            "Information Ratio": 0.15,
            "Maximum_Drawdown": 0.10,
            "Comision": 0.10,
        }

    columnas_requeridas = [
        "Fondo",
        "R2",
        "Beta",
        "Information Ratio",
        "Maximum_Drawdown",
        "Tracking Error Anual",
        columna_comision,
    ]
    faltantes = [c for c in columnas_requeridas if c not in df.columns]
    if faltantes:
        raise ValueError(f"Faltan columnas para calcular el score: {faltantes}")

    # Convertir columnas a numéricas
    for col in [
        "R2",
        "Beta",
        "Information Ratio",
        "Maximum_Drawdown",
        "Tracking Error Anual",
        columna_comision
    ]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    # Normalizar fondo
    fondo_obj_norm = normalizar_fondo(fondo_objetivo)
    df["Fondo_norm"] = df["Fondo"].apply(normalizar_fondo)

    # Validar que exista el fondo objetivo
    df_obj = df[df["Fondo_norm"] == fondo_obj_norm].copy()
    if df_obj.empty:
        raise ValueError(f"No se encontró el fondo objetivo '{fondo_objetivo}' en df_resumen.")

    # Tomar datos del fondo objetivo
    comision_objetivo = df_obj[columna_comision].iloc[0]
    mdd_objetivo = df_obj["Maximum_Drawdown"].iloc[0]

    if pd.isna(comision_objetivo):
        raise ValueError(
            f"La columna '{columna_comision}' del fondo objetivo '{fondo_objetivo}' está vacía o no es numérica."
        )

    if pd.isna(mdd_objetivo):
        raise ValueError(
            f"La columna 'Maximum_Drawdown' del fondo objetivo '{fondo_objetivo}' está vacía o no es numérica."
        )

    # Magnitudes para comparar drawdown
    mdd_objetivo_abs = abs(mdd_objetivo)

    # Máscara de candidatos
    mask_candidatos = df["Fondo_norm"] != fondo_obj_norm

    # Columnas auxiliares para trazabilidad
    # Comisión: positivo = más caro que el objetivo, negativo = más barato
    df["Diferencia_Comision_vs_Objetivo"] = df[columna_comision] - comision_objetivo

    # MDD: positivo = peor drawdown que el objetivo, negativo = mejor drawdown
    df["Diferencia_MDD_vs_Objetivo"] = df["Maximum_Drawdown"].abs() - mdd_objetivo_abs

    # ========= Scores absolutos por métrica =========

    # 1) R2: entre más cerca de 1, mejor
    df["score_R2"] = np.clip(df["R2"] * 100, 0, 100)

    # 2) Tracking Error: entre menor, mejor
    df["score_TE"] = np.clip(
        100 * (1 - (df["Tracking Error Anual"] / umbral_te)),
        0,
        100
    )

    # 3) Beta: lo ideal es 1
    df["score_Beta"] = np.clip(
        100 * (1 - (df["Beta"] - 1).abs() / tolerancia_beta),
        0,
        100
    )

    # 4) Information Ratio: entre más alto, mejor
    df["score_IR"] = np.clip(
        (df["Information Ratio"] / umbral_ir) * 100,
        0,
        100
    )

    # 5) Maximum Drawdown vs fondo objetivo
    # - Si el candidato tiene drawdown <= objetivo (en magnitud): 100
    # - Si tiene drawdown peor: penalización proporcional
    eps = 1e-12
    mdd_candidato_abs = df["Maximum_Drawdown"].abs()

    if mdd_objetivo_abs <= eps:
        # Caso extremadamente raro: el fondo objetivo nunca tuvo drawdown
        df["score_MDD"] = np.where(mdd_candidato_abs <= eps, 100.0, 0.0)
    else:
        df["score_MDD"] = np.where(
            mdd_candidato_abs <= mdd_objetivo_abs,
            100.0,
            np.clip(100 * (mdd_objetivo_abs / mdd_candidato_abs), 0, 100)
        )

    # 6) Comisión vs fondo objetivo
    # - Si el candidato cobra <= objetivo: 100
    # - Si cobra más: penalización proporcional
    if abs(comision_objetivo) <= eps:
        # Caso raro: el objetivo no cobra comisión
        df["score_Comision"] = np.where(df[columna_comision] <= eps, 100.0, 0.0)
    else:
        df["score_Comision"] = np.where(
            df[columna_comision] <= comision_objetivo,
            100.0,
            np.clip(100 * (comision_objetivo / df[columna_comision]), 0, 100)
        )

    # ========= Score final ponderado =========
    df["Score_Sustitucion"] = (
        df["score_R2"] * pesos["R2"] +
        df["score_TE"] * pesos["Tracking Error Anual"] +
        df["score_Beta"] * pesos["Beta"] +
        df["score_IR"] * pesos["Information Ratio"] +
        df["score_MDD"] * pesos["Maximum_Drawdown"] +
        df["score_Comision"] * pesos["Comision"]
    )

    # El objetivo no compite contra sí mismo
    df.loc[~mask_candidatos, "Score_Sustitucion"] = np.nan

    # Ranking solo para candidatos
    df["Ranking_Sustitucion"] = np.nan
    df.loc[mask_candidatos, "Ranking_Sustitucion"] = (
        df.loc[mask_candidatos, "Score_Sustitucion"]
        .rank(method="dense", ascending=False)
    )

    # ========= Orden final para el Excel =========
    # 1) Fondo objetivo primero
    # 2) Candidatos del score más alto al más bajo
    df["Es_Objetivo"] = (~mask_candidatos).astype(int)

    df = df.sort_values(
        by=["Es_Objetivo", "Score_Sustitucion", "Fondo"],
        ascending=[False, False, True],
        na_position="last",
        kind="mergesort"
    ).drop(columns=["Es_Objetivo"])

    return df

def formatear_excel_metricas(ruta_excel: str, hoja=None):
    """
    Aplica formatos al archivo de Excel:
    - J:N  -> porcentaje con 2 decimales
    - O:P  -> número con 3 decimales
    - Q:S  -> porcentaje con 2 decimales
    - T    -> número con 3 decimales
    """
    wb = load_workbook(ruta_excel)
    ws = wb[hoja] if hoja else wb.active

    # Columnas con formato porcentaje 2 decimales
    columnas_porcentaje = ["J", "K", "L", "M", "N", "Q", "R", "S"]

    # Columnas con formato número 3 decimales
    columnas_numero = ["O", "P", "T"]

    # Aplicar formato de porcentaje
    for col in columnas_porcentaje:
        for cell in ws[col][1:]:  # omite encabezado
            if cell.value is not None:
                cell.number_format = "0.00%"

    # Aplicar formato numérico
    for col in columnas_numero:
        for cell in ws[col][1:]:  # omite encabezado
            if cell.value is not None:
                cell.number_format = "0.000"

    # Opcional: congelar encabezado y activar filtro
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(ruta_excel)
    
def limpiar_candidatos(candidatos, fondo_objetivo):
    fondo_obj_norm = normalizar_fondo(fondo_objetivo)
    candidatos_norm = [normalizar_fondo(c) for c in candidatos]
    return list(dict.fromkeys(f for f in candidatos_norm if f != fondo_obj_norm))

def construir_mapa_colores(fondo_objetivo, candidatos):
    fondo_obj_norm = normalizar_fondo(fondo_objetivo)

    # Paleta segura para candidatos
    paleta_candidatos = [
        "#1f77b4",  # azul
        "#2ca02c",  # verde
        "#9467bd",  # morado
        "#17becf",  # cyan
        "#7f7f7f",  # gris medio
        "#bcbd22",  # oliva
        "#ff7f0e",  # naranja
        "#8c564b",  # café
        "#e377c2",  # rosa
        "#aec7e8",  # azul claro
    ]

    # Si no hay candidatos, solo regresa el objetivo
    if candidatos is None:
        return {fondo_obj_norm: "black"}

    # Limpiar candidatos: quitar duplicados y excluir el objetivo
    candidatos_norm = limpiar_candidatos(candidatos, fondo_objetivo)

    mapa_colores = {fondo_obj_norm: "black"}

    for i, fondo_norm in enumerate(candidatos_norm):
        mapa_colores[fondo_norm] = paleta_candidatos[i % len(paleta_candidatos)]

    return mapa_colores

def graficar_ytd_rendimientos_simples(
    df,
    fondo_objetivo,
    candidatos=None,
    anio=None,
    mapa_colores=None,
    guardar=False,
    carpeta_salida=None

):
    df = df.copy()
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df["Rendimiento"] = pd.to_numeric(df["Rendimiento"], errors="coerce")
    df["Fondo_norm"] = df["Fondo"].apply(normalizar_fondo)

    fondo_obj_norm = normalizar_fondo(fondo_objetivo)
    anio = anio or pd.Timestamp.today().year

    # Limpiar lista de candidatos
    if candidatos is None:
        candidatos_norm = [f for f in df["Fondo_norm"].dropna().unique() if f != fondo_obj_norm]
    else:
        candidatos_norm = limpiar_candidatos(candidatos, fondo_objetivo)

    fondos_grafica = list(dict.fromkeys([fondo_obj_norm] + candidatos_norm))

    dff = df[
        (df["Fondo_norm"].isin(fondos_grafica)) &
        (df["Fecha"].dt.year == anio)
    ].copy()

    if dff.empty:
        raise ValueError(f"No hay datos para graficar rendimientos simples en el año {anio}.")

    dff = dff.dropna(subset=["Fecha", "Rendimiento"]).sort_values(["Fondo_norm", "Fecha"])

    mapa_nombres = dff.groupby("Fondo_norm")["Fondo"].first().to_dict()

    plt.figure(figsize=(13, 7))

    # Orden: primero objetivo, luego candidatos
    orden_fondos = [fondo_obj_norm] + [f for f in candidatos_norm if f in dff["Fondo_norm"].unique()]

    for fondo_norm in orden_fondos:
        grp = dff[dff["Fondo_norm"] == fondo_norm].copy()
        if grp.empty:
            continue

        nombre = mapa_nombres.get(fondo_norm, fondo_norm)
        color_fondo = mapa_colores.get(fondo_norm, "gray") if mapa_colores else None

        y = grp["Rendimiento"] * 100

        if fondo_norm == fondo_obj_norm:
            plt.plot(
                grp["Fecha"],
                y,
                linewidth=2.8,
                color=color_fondo or "black",
                label=f"Objetivo - {nombre}",
                zorder=3
            )
        else:
            plt.plot(
                grp["Fecha"],
                y,
                linewidth=1.5,
                alpha=0.90,
                color=color_fondo,
                label=nombre,
                zorder=2
            )

    plt.axhline(0, color="gray", linestyle="--", linewidth=1)
    plt.title(f"Rendimientos simples YTD {anio}: fondo objetivo vs candidatos")
    plt.xlabel("Fecha")
    plt.ylabel("Rendimiento simple (%)")
    plt.legend(loc="best", fontsize=9)
    plt.grid(True, alpha=0.25)
    plt.tight_layout()
    if guardar and carpeta_salida:
        ruta_img = construir_ruta_grafica(
            carpeta_salida,
            fondo_objetivo,
            f"YTD_Rendimientos_Simples_{anio}"
        )
        plt.savefig(ruta_img, dpi=300, bbox_inches="tight")
    plt.show()


def graficar_ytd_vs_objetivo(df, fondo_objetivo, candidatos=None, anio=None, base_100=False, mapa_colores=None,guardar=False,carpeta_salida=None):
    df = df.copy()
    df["Fecha"] = pd.to_datetime(df["Fecha"])
    df["Rendimiento"] = pd.to_numeric(df["Rendimiento"], errors="coerce")
    df["Fondo_norm"] = df["Fondo"].apply(normalizar_fondo)

    fondo_obj_norm = normalizar_fondo(fondo_objetivo)
    anio = anio or pd.Timestamp.today().year

    if candidatos is None:
        candidatos_norm = [f for f in df["Fondo_norm"].dropna().unique() if f != fondo_obj_norm]
    else:
        candidatos_norm = limpiar_candidatos(candidatos, fondo_objetivo)
        # quitar objetivo y duplicados conservando orden
        candidatos_norm = list(dict.fromkeys(
            f for f in candidatos_norm if f != fondo_obj_norm
        ))

    fondos_grafica = list(dict.fromkeys([fondo_obj_norm] + candidatos_norm))

    dff = df[
        (df["Fondo_norm"].isin(fondos_grafica)) &
        (df["Fecha"].dt.year == anio)
    ].copy()

    if dff.empty:
        raise ValueError(f"No hay datos para graficar en el año {anio}.")

    dff = dff.sort_values(["Fondo_norm", "Fecha"])
    dff["Rendimiento_YTD"] = dff.groupby("Fondo_norm")["Rendimiento"].transform(
        lambda s: (1 + s).cumprod() - 1
    )

    mapa_nombres = dff.groupby("Fondo_norm")["Fondo"].first().to_dict()

    plt.figure(figsize=(13, 7))

    orden_fondos = [fondo_obj_norm] + [f for f in candidatos_norm if f in dff["Fondo_norm"].unique()]

    for fondo_norm in orden_fondos:
        grp = dff[dff["Fondo_norm"] == fondo_norm].copy()
        if grp.empty:
            continue

        nombre = mapa_nombres.get(fondo_norm, fondo_norm)
        color_fondo = mapa_colores.get(fondo_norm, "gray") if mapa_colores else None
        y = (1 + grp["Rendimiento_YTD"]) * 100 if base_100 else grp["Rendimiento_YTD"] * 100

        if fondo_norm == fondo_obj_norm:
            plt.plot(
                grp["Fecha"],
                y,
                linewidth=3.2,
                color=color_fondo or "black",
                label=f"Objetivo - {nombre}"
            )
        else:
            plt.plot(
                grp["Fecha"],
                y,
                linewidth=1.8,
                alpha=0.90,
                color=color_fondo,
                label=nombre
            )

    plt.axhline(100 if base_100 else 0, color="gray", linestyle="--", linewidth=1)
    plt.title(f"Rendimiento YTD {anio}: Fondo objetivo vs candidatos")
    plt.xlabel("Fecha")
    plt.ylabel("Índice base 100" if base_100 else "Rendimiento acumulado (%)")
    plt.legend(loc="best", fontsize=9)
    plt.grid(True, alpha=0.25)
    plt.tight_layout()
    if guardar and carpeta_salida:
        nombre = f"YTD_Acumulado_{anio}_Base100" if base_100 else f"YTD_Acumulado_{anio}"
        ruta_img = construir_ruta_grafica(
            carpeta_salida,
            fondo_objetivo,
            nombre
        )
        plt.savefig(ruta_img, dpi=300, bbox_inches="tight")
    plt.show()


def graficar_rendimiento_acumulado(df, fondo_objetivo, candidatos=None, base_100=False, mapa_colores=None,guardar=False,carpeta_salida=None):
    df = df.copy()
    df["Fecha"] = pd.to_datetime(df["Fecha"])
    df["Rendimiento"] = pd.to_numeric(df["Rendimiento"], errors="coerce")
    df["Fondo_norm"] = df["Fondo"].apply(normalizar_fondo)

    fondo_obj_norm = normalizar_fondo(fondo_objetivo)

    if candidatos is None:
        candidatos_norm = [f for f in df["Fondo_norm"].dropna().unique() if f != fondo_obj_norm]
    else:
        candidatos_norm = limpiar_candidatos(candidatos, fondo_objetivo)
        
    fondos_grafica = list(dict.fromkeys([fondo_obj_norm] + candidatos_norm))
    dff = df[df["Fondo_norm"].isin(fondos_grafica)].copy()

    if dff.empty:
        raise ValueError("No hay datos para graficar.")

    dff = dff.dropna(subset=["Rendimiento"]).sort_values(["Fondo_norm", "Fecha"])

    dff["Rendimiento_Acumulado"] = dff.groupby("Fondo_norm")["Rendimiento"].transform(
        lambda s: (1 + s).cumprod() - 1
    )

    mapa_nombres = dff.groupby("Fondo_norm")["Fondo"].first().to_dict()

    plt.figure(figsize=(14, 8))

    orden_fondos = [fondo_obj_norm] + [f for f in candidatos_norm if f in dff["Fondo_norm"].unique()]

    for fondo_norm in orden_fondos:
        grp = dff[dff["Fondo_norm"] == fondo_norm].copy()
        if grp.empty:
            continue

        nombre = mapa_nombres.get(fondo_norm, fondo_norm)
        color_fondo = mapa_colores.get(fondo_norm, "gray") if mapa_colores else None

        y = (1 + grp["Rendimiento_Acumulado"]) * 100 if base_100 else grp["Rendimiento_Acumulado"] * 100

        if fondo_norm == fondo_obj_norm:
            plt.plot(
                grp["Fecha"],
                y,
                linewidth=3.2,
                color=color_fondo or "black",
                label=f"Objetivo - {nombre}",
                zorder=3
            )
        else:
            plt.plot(
                grp["Fecha"],
                y,
                linewidth=1.8,
                alpha=0.95,
                color=color_fondo,
                label=nombre,
                zorder=2
            )

    plt.axhline(100 if base_100 else 0, color="gray", linestyle="--", linewidth=1)
    plt.title("Rendimiento acumulado histórico: fondo objetivo vs candidatos")
    plt.xlabel("Fecha")
    plt.ylabel("Índice base 100" if base_100 else "Rendimiento acumulado (%)")
    plt.legend(loc="best", fontsize=9)
    plt.grid(True, alpha=0.25)
    plt.tight_layout()
    
    if guardar and carpeta_salida:
        nombre = "Rendimiento_Acumulado_Base100" if base_100 else "Rendimiento_Acumulado"
        ruta_img = construir_ruta_grafica(
            carpeta_salida,
            fondo_objetivo,
            nombre
        )
        plt.savefig(ruta_img, dpi=300, bbox_inches="tight")

    plt.show()


def calcular_rendimientos_mensuales_completos(df, fondo_objetivo, candidatos=None):
    df = df.copy()
    df["Fecha"] = pd.to_datetime(df["Fecha"])
    df["Precio"] = pd.to_numeric(df["Precio"], errors="coerce")
    df["Fondo_norm"] = df["Fondo"].apply(normalizar_fondo)

    fondo_obj_norm = normalizar_fondo(fondo_objetivo)

    if candidatos is None:
        candidatos_norm = [f for f in df["Fondo_norm"].dropna().unique() if f != fondo_obj_norm]
    else:
        candidatos_norm = limpiar_candidatos(candidatos, fondo_objetivo)

    fondos = list(dict.fromkeys([fondo_obj_norm] + candidatos_norm))

    dff = df[
        df["Fondo_norm"].isin(fondos)
    ].dropna(subset=["Fecha", "Precio"]).copy()

    if dff.empty:
        raise ValueError("No hay datos para calcular rendimientos mensuales.")

    dff = dff.sort_values(["Fecha", "Fondo_norm"]).copy()
    dff["Mes"] = dff["Fecha"].dt.to_period("M")

    # Fechas de referencia del dataset por mes
    # (primer y último día hábil observado en ese mes)
    bordes_mes = dff.groupby("Mes")["Fecha"].agg(
        FechaInicioMes="min",
        FechaFinMes="max"
    ).reset_index()

    # Resumen por fondo y mes
    mensual = dff.groupby(["Fondo_norm", "Fondo", "Mes"]).agg(
        FechaPrimera=("Fecha", "min"),
        FechaUltima=("Fecha", "max"),
        PrecioInicial=("Precio", "first"),
        PrecioFinal=("Precio", "last")
    ).reset_index()

    mensual = mensual.merge(bordes_mes, on="Mes", how="left")

    # Mes completo = cubre desde el primer hasta el último día hábil observado del mes
    mensual["MesCompleto"] = (
        (mensual["FechaPrimera"] == mensual["FechaInicioMes"]) &
        (mensual["FechaUltima"] == mensual["FechaFinMes"])
    )

    # Conservar solo meses completos para TODOS los fondos
    conteo_completo = mensual.groupby("Mes").agg(
        FondosCompletos=("MesCompleto", "sum"),
        FondosEsperados=("Fondo_norm", "nunique")
    ).reset_index()

    meses_validos = conteo_completo.loc[
        conteo_completo["FondosCompletos"] == conteo_completo["FondosEsperados"],
        "Mes"
    ]

    mensual = mensual[
        (mensual["Mes"].isin(meses_validos)) &
        (mensual["MesCompleto"])
    ].copy()

    if mensual.empty:
        raise ValueError("No hay meses completos comunes entre el fondo objetivo y los candidatos.")

    mensual["Rendimiento_Mensual"] = (mensual["PrecioFinal"] / mensual["PrecioInicial"]) - 1
    mensual["MesEtiqueta"] = mensual["Mes"].astype(str)

    return mensual.sort_values(["Mes", "Fondo_norm"]).reset_index(drop=True)


def graficar_barras_rendimientos_mensuales(df, fondo_objetivo, candidatos=None, mapa_colores=None,guardar=False,carpeta_salida=None):
    mensual = calcular_rendimientos_mensuales_completos(df, fondo_objetivo, candidatos)
    fondo_obj_norm = normalizar_fondo(fondo_objetivo)

    piv = mensual.pivot(index="MesEtiqueta", columns="Fondo", values="Rendimiento_Mensual")

    colores = []
    for col in piv.columns:
        col_norm = normalizar_fondo(col)
        color_fondo = mapa_colores.get(col_norm, "gray") if mapa_colores else None
        colores.append(color_fondo)

    ax = piv.mul(100).plot(
        kind="bar",
        figsize=(15, 8),
        width=0.85,
        color=colores
    )

    # Reforzar visualmente el objetivo
    for patch_group, col in zip(ax.containers, piv.columns):
        if normalizar_fondo(col) == fondo_obj_norm:
            for bar in patch_group:
                bar.set_linewidth(1.3)
                bar.set_edgecolor("black")

    plt.title("Rendimientos mensuales (meses completos): fondo objetivo vs candidatos")
    plt.xlabel("Mes")
    plt.ylabel("Rendimiento mensual (%)")
    plt.axhline(0, color="gray", linewidth=1)
    plt.xticks(rotation=45, ha="right")
    plt.legend(title="Fondo", fontsize=9)
    plt.tight_layout()
    
    if guardar and carpeta_salida:
        ruta_img = construir_ruta_grafica(
            carpeta_salida,
            fondo_objetivo,
            "Rendimientos_Mensuales"
        )
        plt.savefig(ruta_img, dpi=300, bbox_inches="tight")

    plt.show()

    return mensual

def limpiar_nombre_archivo(texto):
    """
    Limpia texto para usarlo como nombre de archivo en Windows.
    """
    texto = str(texto).strip()
    texto = re.sub(r'[<>:"/\\|?*]', '', texto)   # quita caracteres inválidos
    texto = texto.replace(" ", "_")
    return texto


def construir_ruta_grafica(carpeta_base, fondo_objetivo, nombre_grafica, extension="png"):
    """
    Construye la ruta completa del archivo con:
    FondoObjetivo_NombreGrafica.png
    """
    carpeta = Path(carpeta_base)
    carpeta.mkdir(parents=True, exist_ok=True)

    fondo_limpio = limpiar_nombre_archivo(fondo_objetivo)
    grafica_limpia = limpiar_nombre_archivo(nombre_grafica)

    nombre_archivo = f"{fondo_limpio}_{grafica_limpia}.{extension}"
    return carpeta / nombre_archivo

#=========================================================================================================================================

# 1. Obtener el fondo objetivo y los candidatos a través de la lectura de un archivo de Excel
fondo_objetivo,candidatos,error=leer_fondos_excel(r"\\mxmefs02\Mesa de Operaciones\AAAA DENILSON\Fondo Estratega\Objetivo\Objetivo.xlsx")

# Si hay problemas al leer el archivo se lanza un error y se detiene el programa
if error:
    raise ValueError(error)

print("Fondo objetivo:", fondo_objetivo)
print("Candidatos:", candidatos)

fondos = [fondo_objetivo] + candidatos

# 2. Extracción de precios y parametrización de los fondos seleccionados

# Limpia cada nombre quitando espacios y retorna una cadena lista para la lectura en SQL ('FONDO1','FONDO2',...)
lista_sql = formatear_lista_sql(fondos)

#'2023-12-28'

Consulta_sql_precios = f"""
DECLARE @FechaInicioBase DATE = '2023-12-28';
DECLARE @FechaFinObjetivo DATE = '2026-03-31';

;WITH Dias_Inhabiles AS (
    SELECT CAST(FC.Fecha_Calendario_Fecha AS date) AS Fecha
    FROM Fecha_Calendario AS FC
    WHERE FC.Tipo_Calendario_Id = 5
      AND FC.Fecha_Calendario_Fecha > @FechaInicioBase
      AND FC.Fecha_Calendario_Habil = 0
),
PreciosBase AS (
    SELECT 
        PF.Emision_Id,
        PF.Precio_Fondo_Importe,
        CAST(PF.Precio_Fondo_Fecha AS date) AS Fecha
    FROM Precio_Fondo AS PF
    INNER JOIN Fondo 
        ON Fondo.Emision_Id = PF.Emision_Id
    LEFT JOIN Dias_Inhabiles DI 
        ON DI.Fecha = CAST(PF.Precio_Fondo_Fecha AS date)
    WHERE REPLACE(Fondo.Emision_Id,' ','') IN ({lista_sql})
      AND PF.Precio_Fondo_Fecha > @FechaInicioBase
      AND DI.Fecha IS NULL
),
RangoPorFondo AS (
    SELECT
        Emision_Id,
        MIN(Fecha) AS FechaPrimera,
        MAX(Fecha) AS FechaUltima
    FROM PreciosBase
    GROUP BY Emision_Id
),
Limites AS (
    SELECT
        MAX(FechaPrimera) AS FechaInicioComun,
        MIN(FechaUltima) AS FechaFinComunReal
    FROM RangoPorFondo
),
LimitesEfectivos AS (
    SELECT
        FechaInicioComun,
        CASE 
            WHEN FechaFinComunReal >= @FechaFinObjetivo THEN @FechaFinObjetivo
            ELSE FechaFinComunReal
        END AS FechaFinEfectiva
    FROM Limites
),
PreciosFiltrados AS (
    SELECT pb.*
    FROM PreciosBase pb
    CROSS JOIN LimitesEfectivos le
    WHERE pb.Fecha >= le.FechaInicioComun
      AND pb.Fecha <= le.FechaFinEfectiva
)
SELECT
    Fecha,
    Emision_Id AS Fondo,
    Precio_Fondo_Importe as Precio
FROM PreciosFiltrados
ORDER BY Emision_Id, Fecha DESC;
"""
Consulta_sql_Fondos="""
SELECT 
	Fondo.Emision_Id AS Fondo,
	TP.Tipo_Valor_Dsc_Corta AS 'Tipo Valor',
	CASE Fondo.Fondo_Opera_Precio_Cierre 
	WHEN 0 THEN 'NO'
	ELSE 'SÍ'
	END AS 'Opera Precio Cierre',
	CF.Categoria_Fondo_Dsc AS 'Clasificación',
	TF.Tipo_Fondo_Dsc_Corta AS 'Tipo',
	FC.Fondo_Calificacion_Dsc AS 'Calificación',
	PF.Parametros_Fondo_Comision_Distribucion AS 'Comisión de Distribución',
	Fondo.Fondo_Comision_Administracion AS 'Comisión de Administración',
	PF.Parametros_Fondo_Comision_Distribucion*Fondo.Fondo_Comision_Administracion/100 AS 'Comisión Total'
FROM
	Fondo
	LEFT JOIN Parametros_Fondo AS PF
		ON PF.Fondo_Id=Fondo.Fondo_Id
	LEFT JOIN Tipo_Inversionista as TI
		ON TI.Tipo_Inversionista_Id=Pf.Tipo_Inversionista_Id
	LEFT JOIN Categoria_Fondo AS CF
		ON PF.Categoria_Fondo_Id=CF.Categoria_Fondo_Id
	LEFT JOIN Fondo_Calificacion AS FC
		ON PF.Fondo_Calificacion_Id_Mes_Actual=FC.Fondo_Calificacion_Id
	LEFT JOIN Emision
		ON Emision.Emision_Id=Fondo.Emision_Id
	LEFT JOIN Casfim
		ON Casfim.Contrato_Id=Fondo.Contrato_Id
	LEft JOIN Fondo_InstDeposito AS ID
		ON Fondo.Fondo_Id=ID.Fondo_Id
	LEFT JOIN Contrato
		ON contrato.Contrato_Id=ID.Contrato_Id
	LEFT JOIN Fondo_Constitucion AS FCons
		ON FCons.Fondo_Id=Fondo.Fondo_Id
	INNER JOIN Tipo_Fondo AS TF
		ON FCons.Tipo_Fondo_Id=TF.Tipo_Fondo_Id
	INNER JOIN Tipo_Valor AS TP
		ON TP.Tipo_Valor_Id=FCons.Tipo_Valor_Id_Compra


WHERE Fondo.Fondo_Sts_Baja=0
AND Fondo.Emision_Id NOT LIKE'GLREITS%'
AND Emision.Emision_Serie<>'A'
"""

parametros_fondos=Consulta("mxmebas01", "SIIFSK01", Consulta_sql_Fondos)

df = Consulta("mxmebas01", "SIIFSK01", Consulta_sql_precios)

# En caso de que que los data frames retornen vacíos o hubiera un problema en la consulta de sql se detiene el programa
if df is None or df.empty:
    raise ValueError("La consulta de precios no devolvió información.")

if parametros_fondos is None or parametros_fondos.empty:
    raise ValueError("La consulta de parámetros de fondos no devolvió información.")

# 3. Métricas
df = df.sort_values(["Fondo", "Fecha"])

df["Rendimiento"] = df.groupby("Fondo")["Precio"].pct_change()

df = df.dropna(subset=["Rendimiento"])

df["Fecha"] = pd.to_datetime(df["Fecha"])

print("Fecha más antigua para el análisis:", df["Fecha"].min())
    
df_metricas = calcular_resumen_fondos(df, fondo_objetivo)

df_resumen=pd.merge(parametros_fondos,df_metricas,how='inner',on='Fondo')

pesos_base = {
    "R2": 0.30,
    "Tracking Error Anual": 0.05,
    "Beta": 0.10,
    "Information Ratio": 0.30,
    "Maximum_Drawdown": 0.15,
    "Comision": 0.10,
}

df_resumen = agregar_score_sustitucion_absoluto(
    df_resumen,
    fondo_objetivo=fondo_objetivo,
    pesos=pesos_base,
    columna_comision="Comisión de Administración",   # o "Comisión Total"
    umbral_te=0.10,
    umbral_ir=1.0,
    tolerancia_beta=0.30
)

ruta_salida = r"D:\SotDen\Downloads\Metricas.xlsx"

df_resumen.to_excel(ruta_salida, index=False)
formatear_excel_metricas(ruta_salida)

# 4. Gráficas

mapa_colores = construir_mapa_colores(fondo_objetivo, candidatos)

carpeta_descargas = r"D:\SotDen\Downloads"

graficar_ytd_rendimientos_simples(
    df,
    fondo_objetivo,
    candidatos=candidatos,
    anio=2026,
    mapa_colores=mapa_colores,
    guardar=True,
    carpeta_salida=carpeta_descargas
)

graficar_ytd_vs_objetivo(
    df,
    fondo_objetivo,
    candidatos=candidatos,
    mapa_colores=mapa_colores,
    guardar=True,
    carpeta_salida=carpeta_descargas
)

graficar_rendimiento_acumulado(
    df,
    fondo_objetivo,
    candidatos=candidatos,
    base_100=True,
    mapa_colores=mapa_colores,
    guardar=True,
    carpeta_salida=carpeta_descargas
)

df_mensual = graficar_barras_rendimientos_mensuales(
    df,
    fondo_objetivo,
    candidatos=candidatos,
    mapa_colores=mapa_colores,
    guardar=True,
    carpeta_salida=carpeta_descargas
)


#df_mensual.to_excel(r"D:\SotDen\Downloads\Rendimientos_Mensuales.xlsx", index=False)



